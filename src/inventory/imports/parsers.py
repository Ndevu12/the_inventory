"""File parsing utilities for CSV and Excel imports.

Both parsers return the same structure: a list of dicts where keys
are the header column names (stripped and lowercased).
"""

import csv
import io


def parse_csv(file_obj) -> list[dict]:
    """Parse an uploaded CSV file into a list of row dicts.

    Handles both ``InMemoryUploadedFile`` and regular file objects.
    Keys are stripped and lowercased.
    """
    if hasattr(file_obj, "read"):
        content = file_obj.read()
        if isinstance(content, bytes):
            content = content.decode("utf-8-sig")
    else:
        content = file_obj

    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
    return rows


def parse_excel(file_obj) -> list[dict]:
    """Parse an uploaded Excel (.xlsx) file into a list of row dicts.

    Reads the first worksheet.  Keys are stripped and lowercased.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        return []

    headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(raw_headers)]
    rows = []
    for row_values in rows_iter:
        if all(v is None for v in row_values):
            continue
        row = {}
        for h, v in zip(headers, row_values):
            row[h] = str(v).strip() if v is not None else ""
        rows.append(row)

    wb.close()
    return rows
